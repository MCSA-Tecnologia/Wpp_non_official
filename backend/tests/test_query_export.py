from __future__ import annotations

import io
import sys
from pathlib import Path
from types import SimpleNamespace

import pandas as pd
from openpyxl import load_workbook

from app.services import query_export
from app.services.query_export import (
    EXPECTED_COLUMNS,
    _canonicalize_columns,
    _read_only_query,
    _write_workbook,
)
from app.services.source_database import SourceDatabaseCredentials


def test_configured_query_with_leading_comments_is_read_only():
    query = Path(__file__).parents[1] / "app" / "sql" / "contact_export.sql"

    content = query.read_text(encoding="utf-8")

    assert _read_only_query(content)
    assert not _read_only_query("SELECT 1; UPDATE Clientes SET ativo = 0")


def test_export_normalizes_query_aliases_and_formats_the_workbook():
    source = pd.DataFrame(
        [
            {
                "pessoaId": 123,
                "email": "ana@example.com",
                "telefone": "31999999999",
                "credor": "Credor A",
                "campanha": "Campanha A",
                "nome": "Ana Silva",
            }
        ]
    )

    frame = _canonicalize_columns(source)
    content = _write_workbook(frame)
    workbook = load_workbook(io.BytesIO(content))
    sheet = workbook["Contatos"]

    assert list(frame.columns[: len(EXPECTED_COLUMNS)]) == EXPECTED_COLUMNS
    assert frame.loc[0, "Nome"] == "Ana Silva"
    assert frame.loc[0, "Telefone"] == "31999999999"
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:G2"
    assert [cell.value for cell in sheet[1]] == EXPECTED_COLUMNS
    assert sheet["A1"].font.bold
    assert sheet["A1"].fill.fgColor.rgb.endswith("176B4D")


def test_export_uses_saved_credentials_and_returns_xlsx(monkeypatch, tmp_path):
    query_path = tmp_path / "contacts.sql"
    query_path.write_text("SELECT 1", encoding="utf-8")
    settings = SimpleNamespace(
        source_query_path=query_path,
        source_sql_driver="ODBC Driver 18 for SQL Server",
        source_query_timeout_seconds=30,
        source_query_max_rows=10_000,
    )
    credentials = SourceDatabaseCredentials(
        server="sql.example",
        database="Candiotto_STD",
        username="readonly",
        password="secret",
    )
    captured = {}

    class Connection:
        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, traceback):
            return False

    def connect(connection_string, timeout):
        captured["connection_string"] = connection_string
        captured["timeout"] = timeout
        return Connection()

    monkeypatch.setattr(query_export, "get_settings", lambda: settings)
    monkeypatch.setattr(query_export, "source_database_credentials", lambda db: credentials)
    monkeypatch.setattr(
        query_export.pd,
        "read_sql_query",
        lambda sql, connection: pd.DataFrame(
            [{"pessoaId": 1, "telefone": "31999999999", "nome": "Ana"}]
        ),
    )
    monkeypatch.setitem(sys.modules, "pyodbc", SimpleNamespace(connect=connect))

    content = query_export.export_contacts_xlsx(None)
    workbook = load_workbook(io.BytesIO(content))

    assert workbook["Contatos"]["A2"].value == "1"
    assert "SERVER={sql.example}" in captured["connection_string"]
    assert "DATABASE={Candiotto_STD}" in captured["connection_string"]
    assert "UID={readonly}" in captured["connection_string"]
    assert "PWD={secret}" in captured["connection_string"]
    assert captured["timeout"] == 30

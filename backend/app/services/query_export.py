from __future__ import annotations

import io
import re

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from ..config import get_settings
from .source_database import source_database_credentials


EXPECTED_COLUMNS = ["pessoaId", "Nome", "email", "Telefone", "observacao", "Credor", "Campanha"]


def _read_only_query(query: str) -> bool:
    without_comments = re.sub(r"/\*.*?\*/|--[^\r\n]*", " ", query, flags=re.DOTALL)
    normalized = re.sub(r"\s+", " ", without_comments).strip().lower()
    if not normalized.startswith(("select ", "with ")):
        return False
    forbidden = re.compile(
        r"\b(insert|update|delete|drop|alter|create|merge|truncate|execute|exec)\b",
        re.IGNORECASE,
    )
    return forbidden.search(normalized) is None


def _odbc_value(value: str) -> str:
    return "{" + value.replace("}", "}}") + "}"


def _canonicalize_columns(frame: pd.DataFrame) -> pd.DataFrame:
    available = {str(column).casefold(): str(column) for column in frame.columns}
    rename = {
        available[column.casefold()]: column
        for column in EXPECTED_COLUMNS
        if column not in frame.columns and column.casefold() in available
    }
    frame = frame.rename(columns=rename)
    for column in EXPECTED_COLUMNS:
        if column not in frame.columns:
            frame[column] = ""
    leading = EXPECTED_COLUMNS + [
        column for column in frame.columns if column not in EXPECTED_COLUMNS
    ]
    return frame[leading]


def _write_workbook(frame: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        frame.to_excel(writer, index=False, sheet_name="Contatos")
        sheet = writer.sheets["Contatos"]
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        header_fill = PatternFill("solid", fgColor="176B4D")
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
        for index, column in enumerate(frame.columns, start=1):
            values = [str(column), *frame[column].fillna("").astype(str).tolist()]
            sheet.column_dimensions[get_column_letter(index)].width = min(
                max(len(value) for value in values) + 2, 48
            )
    return output.getvalue()


def export_contacts_xlsx(db: Session) -> bytes:
    settings = get_settings()
    credentials = source_database_credentials(db)
    query_path = settings.source_query_path
    if not query_path.exists():
        raise RuntimeError(f"Arquivo de query não encontrado: {query_path}")
    query = query_path.read_text(encoding="utf-8").strip()
    if not _read_only_query(query):
        raise RuntimeError("A query configurada precisa ser somente leitura.")

    try:
        import pyodbc
    except ImportError as exc:
        raise RuntimeError("Instale o extra sqlserver para habilitar a exportação.") from exc

    connection_string = (
        f"DRIVER={_odbc_value(settings.source_sql_driver)};"
        f"SERVER={_odbc_value(credentials.server)};"
        f"DATABASE={_odbc_value(credentials.database)};"
        f"UID={_odbc_value(credentials.username)};PWD={_odbc_value(credentials.password)};"
        "TrustServerCertificate=yes;"
    )
    try:
        with pyodbc.connect(
            connection_string, timeout=settings.source_query_timeout_seconds
        ) as conn:
            frame = pd.read_sql_query(query, conn)
    except Exception as exc:
        raise RuntimeError(
            "Falha ao consultar o SQL Server. Confira as credenciais e a conectividade."
        ) from exc

    frame = _canonicalize_columns(frame.head(settings.source_query_max_rows))
    for column in ("pessoaId", "Telefone"):
        frame[column] = frame[column].map(lambda value: "" if pd.isna(value) else str(value))
    return _write_workbook(frame)
